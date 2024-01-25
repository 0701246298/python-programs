import tkinter as tk
from os.path import isfile

import openpyxl as openpyxl


def get_details():
    global wb
    product_name = nameField.get()
    product_descripion = descField.get("1.0", tk.END)
    product_price = priceField.get()
    details={"name":product_name,
             "desc":product_descripion,
             "price":product_price,
             }
    print(details)

    if isfile('my_workbook.xlsx'):
        pass
    else:

         wb = openpyxl.Workbook()
         ws = wb.create_sheet("Sheet1")
         ws.append(['product Name', 'Product Description', 'Product price'])
         wb.save("my_workbook.xlsv")

wb.openpyxl.load_workbook("my_workbook.xlsx")
ws = wb['sheet1']
ws.append([product_name, product_descripion, product_price])
wb.save("my_workbook.xlsx")

cell=ws['A1']
value=cell.value
print(value)



window = tk.Tk()
window.geometry('500x500')
label = tk.Label(window, text="Product name", bg='yellow')
label.pack()

nameField=tk.Entry(window)
nameField.pack()

label2 = tk.Label(window, text="product description", bg='blue')
label.pack()
descField=tk.Entry(window)
label2.pack()

descField=tk.Text(window,height=10)
descField.pack()


label3 = tk.Label(window, text="product price", bg='white')
label3.pack()

priceField=tk.Entry(window)
priceField.pack()

button = tk.Button(window, text="Save",command=get_details)
button.pack()

# start the event loop
window.mainloop()
